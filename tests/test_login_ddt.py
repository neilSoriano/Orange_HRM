import openpyxl
import pytest
from collections import defaultdict

from pageObjects.LoginPage import LoginPage
from utilities.BaseClass import BaseClass


class TestLoginDDT(BaseClass):
    failed_logins = defaultdict(list)
    valid_logins = defaultdict(list)

    @pytest.mark.regression
    def test_login_ddt(self):
        log = self.get_logger()
        log.info("Testing different credentials using DDT via Excel")

        login = LoginPage(self.driver)
        book = openpyxl.load_workbook(login.book_path)
        sheet = book.active

        for i in range(2, sheet.max_row+1):
            log.info("Entering username...")
            login.set_username(sheet.cell(row=i, column=2).value)
            log.info("Entering password...")
            login.set_password(sheet.cell(row=i, column=3).value)
            login.click_login()

            if self.driver.current_url == self.baseURL:
                log.error("Login failed, adding status to sheet")
                assert "Invalid credentials" in login.get_invalid_text()
                sheet.cell(row=i, column=5).value = "Login failed"
                self.failed_logins[sheet.cell(row=1, column=2).value].append(sheet.cell(row=i, column=2).value)
                self.failed_logins[sheet.cell(row=1, column=3).value].append(sheet.cell(row=i, column=3).value)

            elif self.driver.current_url == login.expected_url:
                log.info("Login passed, adding status to sheet")
                sheet.cell(row=i, column=5).value = "Login passed"
                self.valid_logins[sheet.cell(row=1, column=2).value].append(sheet.cell(row=i, column=2).value)
                self.valid_logins[sheet.cell(row=1, column=3).value].append(sheet.cell(row=i, column=3).value)

        print(self.failed_logins)
        print("\n")
        print(self.valid_logins)

        book.save(login.book_path)
        book.close()
        log.info("End of DDT")






















